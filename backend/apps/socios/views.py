import io
import json
import calendar
import uuid
import math
from decimal import Decimal
from datetime import datetime, date

from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db import connection, transaction
from django.db.models import Q, Count, Prefetch
from django.db.models.functions import Lower, Trim
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from drf_spectacular.utils import extend_schema, OpenApiResponse
from rest_framework import permissions, status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from .audit import snapshot_socio, register_audit_entry
from .models import Prestamo, Socio, SocioAuditLog, TipoPrestamo, PoliticaAprobacion, Desembolso, Pago
from .serializers import (
    HistorialCrediticioSerializer,
    ProfileCreateSerializer,
    SocioAdminUpdateSerializer,
    SocioEstadoSerializer,
    SocioSerializer,
    TipoPrestamoSerializer,
    TipoPrestamoUpsertSerializer,
    PoliticaAprobacionSerializer,
    PoliticaAprobacionUpsertSerializer,
    PrestamoSimulacionSerializer,
    PrestamoSolicitudSerializer,
    DesembolsoSerializer,
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="43A59D")


def style_header_row(ws, row_idx: int = 1):
    """Apply a simple header style to the given row."""
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def wrap_columns(ws, col_letters: list[str]):
    """Enable wrap_text for the provided columns (skip header)."""
    for col in col_letters:
        for cell in ws[col][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def fmt_decimal(valor: Decimal) -> str:
    """Devuelve el decimal con 2 decimales en formato string."""
    return f"{valor.quantize(Decimal('0.01'))}"


def calcular_tabla_amortizacion(monto: Decimal, tasa_anual: Decimal, plazo_meses: int):
    """Calcula cuota, totales e historial simple de amortizacion."""
    if plazo_meses <= 0:
        raise ValidationError({'plazo_meses': 'El plazo en meses debe ser mayor a 0.'})

    tasa_mensual = (Decimal(tasa_anual) / Decimal('100')) / Decimal('12')
    if tasa_mensual > 0:
        factor = (Decimal('1') + tasa_mensual) ** (-plazo_meses)
        cuota = monto * tasa_mensual / (Decimal('1') - factor)
    else:
        cuota = monto / Decimal(plazo_meses)

    cuota = cuota.quantize(Decimal('0.01'))
    saldo = monto
    cuotas = []
    for numero in range(1, plazo_meses + 1):
        interes = (saldo * tasa_mensual).quantize(Decimal('0.01')) if tasa_mensual > 0 else Decimal('0.00')
        capital = (cuota - interes).quantize(Decimal('0.01'))
        saldo = (saldo - capital).quantize(Decimal('0.01'))
        if saldo < Decimal('0'):
            saldo = Decimal('0.00')
        cuotas.append({
            "numero": numero,
            "cuota": fmt_decimal(cuota),
            "capital": fmt_decimal(capital),
            "interes": fmt_decimal(interes),
            "saldo": fmt_decimal(saldo),
        })

    total_a_pagar = cuota * plazo_meses
    total_intereses = total_a_pagar - monto
    return {
        "cuota_mensual": fmt_decimal(cuota),
        "total_a_pagar": fmt_decimal(total_a_pagar),
        "total_intereses": fmt_decimal(total_intereses if total_intereses > Decimal('0') else Decimal('0')),
        "cuotas": cuotas,
    }


def add_months(fecha: date, months: int) -> date:
    """Suma meses a una fecha manteniendo el dia valido."""
    month = fecha.month - 1 + months
    year = fecha.year + month // 12
    month = month % 12 + 1
    day = min(fecha.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def get_table_columns(table_name: str) -> set[str]:
    """Retorna las columnas existentes de una tabla (schema public)."""
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "postgresql":
            cursor.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                """,
                ["public", table_name],
            )
            return {row[0] for row in cursor.fetchall()}
        # SQLite fallback
        cursor.execute(f"PRAGMA table_info({table_name})")
        return {row[1] for row in cursor.fetchall()}


def get_table_metadata(table_name: str):
    """Retorna metadatos simples de columnas: nombre, nullable, default."""
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "postgresql":
            cursor.execute(
                """
                SELECT column_name, is_nullable, column_default, data_type
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                """,
                ["public", table_name],
            )
            return [
                {
                    "name": row[0],
                    "nullable": row[1] == "YES",
                    "has_default": row[2] is not None,
                    "type": row[3],
                }
                for row in cursor.fetchall()
            ]
        cursor.execute(f"PRAGMA table_info({table_name})")
        return [
            {
                "name": row[1],
                "nullable": not bool(row[3]),  # 0 => nullable in sqlite pragma
                "has_default": row[4] is not None,
                "type": row[2],
            }
            for row in cursor.fetchall()
        ]


def ensure_producto_from_tipo(tipo) -> tuple[uuid.UUID | None, str | None]:
    """Garantiza que exista un registro en producto_prestamo con el id del tipo."""
    meta = get_table_metadata("producto_prestamo")
    if not meta:
        return None, "Tabla producto_prestamo no disponible."

    producto_id = tipo.id
    pid_param = str(producto_id)
    table_name = "producto_prestamo" if connection.vendor != "postgresql" else "public.producto_prestamo"
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT 1 FROM {table_name} WHERE id = %s LIMIT 1", [pid_param])
        if cursor.fetchone():
            return producto_id, None

    now = timezone.now()
    base_payload = {
        "id": producto_id,
        "nombre": getattr(tipo, "nombre", None),
        "descripcion": getattr(tipo, "descripcion", "") if hasattr(tipo, "descripcion") else "",
        "tasa_interes": getattr(tipo, "tasa_interes_anual", None),
        "plazo_meses": getattr(tipo, "plazo_meses", None),
        # Campos adicionales comunes en producto_prestamo
        "tipo": getattr(tipo, "nombre", None) or "producto",
        "tasa_nominal_anual": getattr(tipo, "tasa_interes_anual", None),
        "plazo_max_meses": getattr(tipo, "plazo_meses", None),
        "activo": True,
        "created_at": now,
        "updated_at": now,
    }

    columnas_insert: list[str] = []
    valores: list = []
    faltantes: list[str] = []
    for col in meta:
        name = col["name"]
        if name in base_payload and base_payload[name] is not None:
            columnas_insert.append(name)
            valores.append(base_payload[name])
        elif col["nullable"] or col["has_default"]:
            continue
        elif name == "nombre":
            columnas_insert.append(name)
            valores.append(base_payload.get("nombre") or "Producto generado")
        else:
            faltantes.append(name)

    if faltantes:
        return None, f"Faltan columnas obligatorias en producto_prestamo: {', '.join(faltantes)}"

    if not columnas_insert:
        return None, "No hay columnas válidas para insertar en producto_prestamo."

    placeholders = ", ".join(["%s"] * len(columnas_insert))
    columnas_sql = ", ".join(columnas_insert)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"INSERT INTO {table_name} ({columnas_sql}) VALUES ({placeholders})", [pid_param if v == producto_id else v for v in valores])
    except Exception as exc:
        return None, f"No se pudo crear producto_prestamo: {exc}"

    return producto_id, None


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        socio = getattr(request.user, 'socio', None)
        if not socio:
            return Response({'detail': 'Perfil de socio no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(SocioSerializer(socio).data)


class ProfileUpsertView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = ProfileCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        socio = serializer.save()
        return Response(SocioSerializer(socio).data, status=status.HTTP_201_CREATED)


class SocioListView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Socios'],
        responses=SocioSerializer(many=True),
        summary='Listado de socios',
        description='Devuelve todos los socios registrados. Solo administradores.',
    )
    def get(self, request):
        queryset = Socio.objects.select_related('usuario').order_by('nombre_completo')

        q = (request.query_params.get('q') or '').strip()
        if q:
            queryset = queryset.filter(
                Q(nombre_completo__icontains=q)
                | Q(documento__icontains=q)
                | Q(usuario__email__icontains=q)
                | Q(id__icontains=q)
            )

        return Response(SocioSerializer(queryset, many=True).data)


class TipoPrestamoListCreateView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Prestamos'],
        responses=TipoPrestamoSerializer(many=True),
        summary='Listado de tipos de préstamo',
        description='Devuelve los tipos de préstamo configurados. Solo administradores.',
    )
    def get(self, request):
        qs = TipoPrestamo.objects.all().order_by('nombre')
        q = (request.query_params.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))

        solo_activos = request.query_params.get('solo_activos') or request.query_params.get('soloActivos')
        if solo_activos and str(solo_activos).lower() in {'1', 'true', 't', 'yes', 'on'}:
            qs = qs.filter(activo=True)

        return Response(TipoPrestamoSerializer(qs, many=True).data)

    @extend_schema(
        tags=['Prestamos'],
        request=TipoPrestamoUpsertSerializer,
        responses={201: TipoPrestamoSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Crear tipo de préstamo',
        description='Registra un nuevo tipo de préstamo con tasa, plazo y requisitos.',
    )
    def post(self, request):
        serializer = TipoPrestamoUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tipo = serializer.save()
        return Response(TipoPrestamoSerializer(tipo).data, status=status.HTTP_201_CREATED)


class TipoPrestamoDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get_object(self, tipo_id):
        return get_object_or_404(TipoPrestamo, pk=tipo_id)

    @extend_schema(
        tags=['Prestamos'],
        responses={200: TipoPrestamoSerializer, 404: OpenApiResponse(description='Tipo no encontrado')},
        summary='Detalle de tipo de préstamo',
        description='Obtiene la definición completa de un tipo de préstamo.',
    )
    def get(self, _request, tipo_id):
        tipo = self.get_object(tipo_id)
        return Response(TipoPrestamoSerializer(tipo).data)

    @extend_schema(
        tags=['Prestamos'],
        request=TipoPrestamoUpsertSerializer,
        responses={200: TipoPrestamoSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Actualizar tipo de préstamo',
        description='Permite modificar tasa, plazo, requisitos o estado activo.',
    )
    def put(self, request, tipo_id):
        tipo = self.get_object(tipo_id)
        serializer = TipoPrestamoUpsertSerializer(instance=tipo, data=request.data)
        serializer.is_valid(raise_exception=True)
        tipo = serializer.save()
        return Response(TipoPrestamoSerializer(tipo).data)

    @extend_schema(
        tags=['Prestamos'],
        request=TipoPrestamoUpsertSerializer,
        responses={200: TipoPrestamoSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Actualización parcial de tipo de préstamo',
        description='Actualiza parcialmente un tipo de préstamo.',
    )
    def patch(self, request, tipo_id):
        tipo = self.get_object(tipo_id)
        serializer = TipoPrestamoUpsertSerializer(instance=tipo, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        tipo = serializer.save()
        return Response(TipoPrestamoSerializer(tipo).data)

    @extend_schema(
        tags=['Prestamos'],
        responses={200: TipoPrestamoSerializer, 404: OpenApiResponse(description='Tipo no encontrado')},
        summary='Desactivar tipo de préstamo',
        description='Desactiva (soft delete) un tipo de préstamo para que no se use en nuevos créditos.',
    )
    def delete(self, _request, tipo_id):
        tipo = self.get_object(tipo_id)
        if tipo.activo:
            tipo.activo = False
            tipo.save(update_fields=['activo', 'updated_at'])
        return Response(TipoPrestamoSerializer(tipo).data, status=status.HTTP_200_OK)


class TipoPrestamoPublicListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=['Prestamos'],
        responses=TipoPrestamoSerializer(many=True),
        summary='Tipos de prestamo activos (socios)',
        description='Listado simplificado de tipos de prestamo activos para que el socio pueda solicitarlos.',
    )
    def get(self, _request):
        qs = TipoPrestamo.objects.filter(activo=True).order_by('nombre')
        return Response(TipoPrestamoSerializer(qs, many=True).data)


class PrestamoSimulacionView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=['Prestamos'],
        request=PrestamoSimulacionSerializer,
        responses={200: OpenApiResponse(description='Simulacion generada'), 404: OpenApiResponse(description='Socio o tipo no encontrado')},
        summary='Simular prestamo de socio',
        description='Calcula cuota mensual, total a pagar e intereses usando el tipo de prestamo seleccionado.',
    )
    def post(self, request):
        socio = getattr(request.user, 'socio', None)
        if not socio:
            return Response({'detail': 'Perfil de socio no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        if socio.estado != Socio.ESTADO_ACTIVO:
            return Response({'detail': 'El socio no se encuentra activo.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = PrestamoSimulacionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        tipo = get_object_or_404(TipoPrestamo, pk=serializer.validated_data['tipo_prestamo_id'], activo=True)
        monto = serializer.validated_data['monto']
        plazo = serializer.validated_data.get('plazo_meses') or tipo.plazo_meses
        if plazo > tipo.plazo_meses:
            return Response({'plazo_meses': f'Maximo permitido para este producto: {tipo.plazo_meses} meses.'}, status=status.HTTP_400_BAD_REQUEST)
        if plazo < 6 or plazo % 6 != 0:
            return Response({'plazo_meses': 'El plazo debe ser en saltos de 6 meses y al menos 6.'}, status=status.HTTP_400_BAD_REQUEST)

        plan = calcular_tabla_amortizacion(monto, tipo.tasa_interes_anual, plazo)

        data = {
            "socio": {
                "id": str(socio.id),
                "nombre_completo": socio.nombre_completo,
                "documento": socio.documento,
                "email": socio.usuario.email if socio.usuario else None,
            },
            "tipo": TipoPrestamoSerializer(tipo).data,
            "monto": fmt_decimal(monto),
            "plazo_meses": plazo,
            **plan,
        }
        return Response(data, status=status.HTTP_200_OK)


class PrestamoSolicitudCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=['Prestamos'],
        request=PrestamoSolicitudSerializer,
        responses={201: OpenApiResponse(description='Solicitud registrada'), 404: OpenApiResponse(description='Socio o tipo no encontrado')},
        summary='Registrar solicitud de prestamo',
        description='Crea un prestamo asociado al socio autenticado usando un tipo activo y devuelve el resumen calculado.',
    )
    def post(self, request):
        socio = getattr(request.user, 'socio', None)
        if not socio:
            return Response({'detail': 'Perfil de socio no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        if socio.estado != Socio.ESTADO_ACTIVO:
            return Response({'detail': 'El socio no se encuentra activo.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = PrestamoSolicitudSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        tipo = get_object_or_404(TipoPrestamo, pk=serializer.validated_data['tipo_prestamo_id'], activo=True)
        monto = serializer.validated_data['monto']
        plazo = serializer.validated_data.get('plazo_meses') or tipo.plazo_meses
        if plazo > tipo.plazo_meses:
            return Response({'plazo_meses': f'Maximo permitido para este producto: {tipo.plazo_meses} meses.'}, status=status.HTTP_400_BAD_REQUEST)
        if plazo < 6 or plazo % 6 != 0:
            return Response({'plazo_meses': 'El plazo debe ser en saltos de 6 meses y al menos 6.'}, status=status.HTTP_400_BAD_REQUEST)

        plan = calcular_tabla_amortizacion(monto, tipo.tasa_interes_anual, plazo)

        # Insertar en tabla de solicitudes (no crea prestamo aún)
        columnas = get_table_columns('solicitud')
        if not columnas:
            return Response({'detail': 'Tabla de solicitud no disponible en la base de datos.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        solicitud_id = uuid.uuid4()
        ahora = timezone.now()
        descripcion = serializer.validated_data.get('descripcion') or ""
        producto_id = None
        if "producto_id" in columnas:
            producto_id, error_producto = ensure_producto_from_tipo(tipo)
            if error_producto:
                return Response({'detail': error_producto}, status=status.HTTP_400_BAD_REQUEST)
        # Campos candidatos (solo se insertan los que existan en la tabla)
        payload = {
            "id": solicitud_id,
            "socio_id": socio.id,
            "monto": monto,
            "tasa_interes": tipo.tasa_interes_anual,
            "plazo_meses": plazo,
            "descripcion": descripcion,
            "estado": "pendiente",
            "created_at": ahora,
            "updated_at": ahora,
        }
        # producto_id: obligatorio en schema actual, usamos el id del tipo de prestamo
        if producto_id:
            payload["producto_id"] = producto_id
        # tipo_prestamo_id por compatibilidad si la columna existe
        if "tipo_prestamo_id" in columnas:
            payload["tipo_prestamo_id"] = tipo.id
        cols_presentes = [col for col in payload.keys() if col in columnas]
        if not cols_presentes:
            return Response({'detail': 'No hay columnas compatibles para guardar la solicitud.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        placeholders = ", ".join(["%s"] * len(cols_presentes))
        columnas_sql = ", ".join(cols_presentes)
        valores = [payload[col] for col in cols_presentes]
        solicitud_table = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"
        if connection.vendor != "postgresql":
            valores = [str(v) if isinstance(v, uuid.UUID) else v for v in valores]
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    f"INSERT INTO {solicitud_table} ({columnas_sql}) VALUES ({placeholders})",
                    valores,
                )
        except Exception as exc:
            return Response({'detail': f'No se pudo registrar la solicitud: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Notificar a analistas (simple)
        return Response(
            {
                "solicitud_id": str(solicitud_id),
                "estado": payload.get("estado") if "estado" in columnas else "registrada",
                "fecha_desembolso": None,
                "fecha_vencimiento": None,
                "socio": {
                    "id": str(socio.id),
                    "nombre_completo": socio.nombre_completo,
                    "documento": socio.documento,
                    "email": socio.usuario.email if socio.usuario else None,
                },
                "tipo": TipoPrestamoSerializer(tipo).data,
                "monto": fmt_decimal(monto),
                "plazo_meses": tipo.plazo_meses,
                **plan,
            },
            status=status.HTTP_201_CREATED,
        )


class MisPrestamosSocioView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["Prestamos"],
        summary="Mis prestamos (socio)",
        description="Devuelve solicitudes y prestamos del socio autenticado incluyendo estado visible y posibilidad de pago cuando hay desembolso.",
    )
    def get(self, request):
        socio = getattr(request.user, "socio", None)
        if not socio:
            return Response({"detail": "Perfil de socio no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        desembolso_prefetch, desembolso_meta = _desembolso_prefetch()
        prefetches = ["pagos"]
        if desembolso_prefetch:
            prefetches.append(desembolso_prefetch)

        solicitudes = {str(row.get("id")): row for row in _solicitudes_por_socio(socio.id, limit=100)}
        prestamos_qs = (
            Prestamo.objects.filter(socio=socio)
            .select_related("tipo")
            .prefetch_related(*prefetches)
            .order_by("-fecha_desembolso", "-created_at")
        )

        resumen = {
            "pendientes": 0,
            "aprobados": 0,
            "rechazados": 0,
            "desembolsados": 0,
            "pagados": 0,
        }
        resumen_map = {
            "pendiente": "pendientes",
            "aprobado": "aprobados",
            "rechazado": "rechazados",
            "desembolsado": "desembolsados",
            "pagado": "pagados",
        }

        resultados = []
        for prestamo in prestamos_qs:
            solicitud = solicitudes.pop(str(prestamo.id), None)
            pagos = list(prestamo.pagos.all())
            desembolsos = list(prestamo.desembolsos.all()) if desembolso_meta["cols"] else []
            plan_info = _plan_cliente_para_prestamo(prestamo, solicitud)
            total_pagado = sum((p.monto for p in pagos), Decimal("0"))
            saldo = prestamo.monto - total_pagado
            if saldo < Decimal("0"):
                saldo = Decimal("0")
            cuota_dec = plan_info.get("cuota_decimal", Decimal("0"))
            cuotas_restantes = 0
            if cuota_dec > Decimal("0") and saldo > Decimal("0"):
                cuotas_restantes = max(1, math.ceil(saldo / cuota_dec))

            estado_visible = _estado_cliente_prestamo(prestamo, solicitud, bool(desembolsos))
            resumen_key = resumen_map.get(estado_visible)
            if resumen_key:
                resumen[resumen_key] += 1

            resultados.append(
                {
                    "id": str(prestamo.id),
                    "solicitud_id": str(solicitud.get("id") or prestamo.id) if solicitud else str(prestamo.id),
                    "estado": estado_visible,
                    "monto": fmt_decimal(prestamo.monto),
                    "cuota_mensual": plan_info.get("cuota_mensual"),
                    "plazo_meses": plan_info.get("plazo_meses"),
                    "tipo": {
                        "id": str(prestamo.tipo.id) if prestamo.tipo else None,
                        "nombre": prestamo.tipo.nombre if prestamo.tipo else None,
                    },
                    "descripcion": (solicitud or {}).get("descripcion") or prestamo.descripcion,
                    "fecha_solicitud": (solicitud or {}).get("created_at"),
                    "fecha_desembolso": prestamo.fecha_desembolso,
                    "fecha_vencimiento": prestamo.fecha_vencimiento,
                    "total_pagado": fmt_decimal(total_pagado),
                    "saldo_pendiente": fmt_decimal(saldo),
                    "pagos_registrados": len(pagos),
                    "cuotas_restantes": cuotas_restantes,
                    "tiene_desembolso": bool(desembolsos),
                    "puede_pagar": bool(desembolsos) and saldo > Decimal("0"),
                }
            )

        for solicitud in solicitudes.values():
            estado_raw = (solicitud.get("estado") or "pendiente").strip().lower()
            resumen_key = resumen_map.get(estado_raw)
            if resumen_key:
                resumen[resumen_key] += 1
            try:
                monto_dec = Decimal(str(solicitud.get("monto") or "0"))
            except Exception:
                monto_dec = Decimal("0")
            plazo_raw = solicitud.get("plazo_meses")
            try:
                plazo = int(plazo_raw) if plazo_raw is not None else None
            except (TypeError, ValueError):
                plazo = None

            resultados.append(
                {
                    "id": str(_fmt_uuid(solicitud.get("id"))),
                    "solicitud_id": str(_fmt_uuid(solicitud.get("id"))),
                    "estado": estado_raw or "pendiente",
                    "monto": fmt_decimal(monto_dec),
                    "cuota_mensual": None,
                    "plazo_meses": plazo,
                    "tipo": None,
                    "descripcion": solicitud.get("descripcion") or "",
                    "fecha_solicitud": solicitud.get("created_at"),
                    "fecha_desembolso": None,
                    "fecha_vencimiento": None,
                    "total_pagado": "0.00",
                    "saldo_pendiente": fmt_decimal(monto_dec),
                    "pagos_registrados": 0,
                    "cuotas_restantes": plazo or 0,
                    "tiene_desembolso": False,
                    "puede_pagar": False,
                }
            )

        def _parse_fecha(valor):
            if isinstance(valor, datetime):
                return valor
            if isinstance(valor, date):
                return datetime.combine(valor, datetime.min.time())
            if isinstance(valor, str):
                try:
                    return datetime.fromisoformat(valor)
                except Exception:
                    return datetime.min
            return datetime.min

        resultados.sort(
            key=lambda item: _parse_fecha(item.get("fecha_desembolso") or item.get("fecha_solicitud")),
            reverse=True,
        )

        return Response({"prestamos": resultados, "resumen": resumen}, status=status.HTTP_200_OK)


class PagoSimuladoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["Prestamos"],
        summary="Registrar pago simulado de cuotas",
        description="Permite al socio pagar una o varias cuotas de un préstamo desembolsado (pasarela simulada).",
    )
    def post(self, request, prestamo_id: uuid.UUID):
        socio = getattr(request.user, "socio", None)
        if not socio:
            return Response({"detail": "Perfil de socio no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        desembolso_prefetch, desembolso_meta = _desembolso_prefetch()
        prefetches = ["pagos"]
        if desembolso_prefetch:
            prefetches.append(desembolso_prefetch)

        prestamo = (
            Prestamo.objects.filter(pk=prestamo_id, socio=socio)
            .select_related("tipo")
            .prefetch_related(*prefetches)
            .first()
        )
        if not prestamo:
            raise Http404("Prestamo no encontrado.")

        desembolsos = list(prestamo.desembolsos.all()) if desembolso_meta["cols"] else []
        if not desembolsos:
            return Response({"detail": "El prestamo aun no esta desembolsado."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cuotas = int((request.data or {}).get("cuotas") or 0)
        except (TypeError, ValueError):
            return Response({"cuotas": "Ingresa un numero de cuotas valido."}, status=status.HTTP_400_BAD_REQUEST)
        if cuotas < 1:
            return Response({"cuotas": "Debes elegir al menos 1 cuota."}, status=status.HTTP_400_BAD_REQUEST)

        metodo = (request.data or {}).get("metodo") or (request.data or {}).get("metodo_pago") or "pasarela"
        solicitud_rel = None
        try:
            solicitud_rel, _ = _fetch_solicitud_row(prestamo.id)
        except Http404:
            solicitudes = _solicitudes_por_socio(socio.id, limit=120)
            solicitud_rel = next((s for s in solicitudes if str(s.get("id")) == str(prestamo.id)), None)

        plan_info = _plan_cliente_para_prestamo(prestamo, solicitud_rel)
        cuota_dec = plan_info.get("cuota_decimal", Decimal("0"))
        if cuota_dec <= Decimal("0"):
            plazo = plan_info.get("plazo_meses") or cuotas
            try:
                plazo_int = int(plazo) if plazo else cuotas
            except (TypeError, ValueError):
                plazo_int = cuotas
            cuota_dec = (prestamo.monto / Decimal(plazo_int or 1)).quantize(Decimal("0.01"))

        total_pagado_actual = sum((p.monto for p in prestamo.pagos.all()), Decimal("0"))
        saldo = prestamo.monto - total_pagado_actual
        if saldo <= Decimal("0"):
            return Response({"detail": "El prestamo ya no tiene saldo pendiente."}, status=status.HTTP_400_BAD_REQUEST)

        monto_pagar = (cuota_dec * Decimal(cuotas)).quantize(Decimal("0.01"))
        if monto_pagar <= Decimal("0"):
            monto_pagar = saldo
        if monto_pagar > saldo:
            monto_pagar = saldo

        pago = Pago.objects.create(
            prestamo=prestamo,
            monto=monto_pagar,
            fecha_pago=timezone.now().date(),
            metodo=str(metodo)[:50],
            referencia=f"SIM-{timezone.now():%Y%m%d%H%M%S}-{uuid.uuid4().hex[:6].upper()}",
        )

        total_pagado = sum((p.monto for p in prestamo.pagos.all()), Decimal("0"))
        saldo_restante = prestamo.monto - total_pagado
        if saldo_restante < Decimal("0"):
            saldo_restante = Decimal("0")

        if saldo_restante == Decimal("0"):
            prestamo.estado = Prestamo.Estados.PAGADO
            prestamo.save(update_fields=["estado", "updated_at"])
            estado_visible = "pagado"
        else:
            estado_visible = _estado_cliente_prestamo(prestamo, solicitud_rel, bool(desembolsos))

        return Response(
            {
                "pago": {
                    "id": pago.id,
                    "monto": fmt_decimal(pago.monto),
                    "fecha_pago": pago.fecha_pago,
                    "metodo": pago.metodo,
                    "referencia": pago.referencia,
                },
                "prestamo": {
                    "id": str(prestamo.id),
                    "estado": estado_visible,
                    "total_pagado": fmt_decimal(total_pagado),
                    "saldo_pendiente": fmt_decimal(saldo_restante),
                },
            },
            status=status.HTTP_201_CREATED,
        )


def _is_analista(user) -> bool:
    rol_nombre = getattr(getattr(user, "rol", None), "nombre", "") or getattr(user, "rol", "") or ""
    return rol_nombre.upper() == "ANALISTA" or getattr(user, "is_staff", False) or getattr(user, "is_superuser", False)

def _is_tesorero(user) -> bool:
    rol_nombre = getattr(getattr(user, "rol", None), "nombre", "") or getattr(user, "rol", "") or ""
    return rol_nombre.upper() == "TESORERO" or getattr(user, "is_staff", False) or getattr(user, "is_superuser", False)


def _fetch_solicitud_row(solicitud_id: uuid.UUID) -> tuple[dict, set[str]]:
    columnas = get_table_columns("solicitud")
    if not columnas:
        raise Http404("Tabla solicitud no encontrada.")
    posibles = [
        "id",
        "socio_id",
        "monto",
        "tasa_interes",
        "plazo_meses",
        "descripcion",
        "estado",
        "created_at",
        "updated_at",
        "producto_id",
        "tipo_prestamo_id",
    ]
    seleccionadas = [c for c in posibles if c in columnas]
    if "observaciones" in columnas:
        seleccionadas.append("observaciones")
    if not seleccionadas:
        raise Http404("No hay columnas legibles en la tabla solicitud.")

    table_name = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"
    solicitud_str = str(solicitud_id)
    with connection.cursor() as cursor:
        cursor.execute(
            f"SELECT {', '.join(seleccionadas)} FROM {table_name} WHERE id = %s LIMIT 1",
            [solicitud_str],
        )
        row = cursor.fetchone()
    if not row:
        raise Http404("Solicitud no encontrada.")
    data = dict(zip(seleccionadas, row))
    return data, columnas


def _extraer_observaciones(row: dict) -> str:
    if "observaciones" in row:
        return row.get("observaciones") or ""
    if "comentarios" in row:
        return row.get("comentarios") or ""
    descripcion = row.get("descripcion") or ""
    marker = "[OBS_ANALISTA]"
    lines = []
    for line in descripcion.splitlines():
        if line.strip().startswith(marker):
            lines.append(line.replace(marker, "", 1).strip())
    return "\n".join(lines)


def _adjuntar_observacion_en_descripcion(descripcion_actual: str, nueva_obs: str, marker: str = "[OBS_ANALISTA]") -> str:
    base = descripcion_actual or ""
    if marker in base and nueva_obs.strip() in base:
        return base
    separator = "\n" if base else ""
    return f"{base}{separator}{marker} {nueva_obs.strip()}"


def _obs_column(columnas: set[str]) -> str | None:
    if "observaciones" in columnas:
        return "observaciones"
    if "comentarios" in columnas:
        return "comentarios"
    return None


def _solicitudes_por_socio(socio_id: uuid.UUID, limit: int = 50) -> list[dict]:
    """Devuelve las solicitudes registradas por el socio (si existe la tabla)."""
    columnas = get_table_columns("solicitud")
    if not columnas:
        return []
    posibles = [
        "id",
        "socio_id",
        "monto",
        "tasa_interes",
        "plazo_meses",
        "descripcion",
        "estado",
        "created_at",
        "updated_at",
        "producto_id",
        "tipo_prestamo_id",
    ]
    seleccionadas = [c for c in posibles if c in columnas]
    if "observaciones" in columnas:
        seleccionadas.append("observaciones")
    if not seleccionadas:
        return []

    table_name = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"
    socio_param = str(socio_id)
    sql = f"""
        SELECT {', '.join(seleccionadas)}
        FROM {table_name}
        WHERE socio_id = %s
        ORDER BY created_at DESC
        LIMIT %s
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [socio_param, limit])
        rows = cursor.fetchall()

    data = []
    for row in rows:
        record = dict(zip(seleccionadas, row))
        record["id"] = _fmt_uuid(record.get("id"))
        record["socio_id"] = _fmt_uuid(record.get("socio_id"))
        data.append(record)
    return data


def _estado_cliente_prestamo(prestamo: Prestamo, solicitud: dict | None, tiene_desembolso: bool) -> str:
    estado_bruto = (prestamo.estado or (solicitud or {}).get("estado") or "").strip().lower()
    if tiene_desembolso:
        return "desembolsado"
    if estado_bruto in {"pendiente", "rechazado"}:
        return estado_bruto
    if estado_bruto in {"aprobado", Prestamo.Estados.ACTIVO, "activo"}:
        return "aprobado"
    if estado_bruto == Prestamo.Estados.PAGADO:
        return "pagado"
    if estado_bruto == Prestamo.Estados.MOROSO:
        return "moroso"
    if estado_bruto == Prestamo.Estados.CANCELADO:
        return "cancelado"
    return estado_bruto or "pendiente"


def _plan_cliente_para_prestamo(prestamo: Prestamo, solicitud: dict | None) -> dict:
    """Calcula cuota mensual estimada para el prestamo con data de solicitud o tipo."""
    plazo_raw = solicitud.get("plazo_meses") if solicitud else None
    try:
        plazo_meses = int(plazo_raw) if plazo_raw is not None else None
    except (TypeError, ValueError):
        plazo_meses = None
    if not plazo_meses or plazo_meses <= 0:
        plazo_meses = getattr(prestamo.tipo, "plazo_meses", None) or 12

    tasa_raw = None
    if solicitud:
        tasa_raw = solicitud.get("tasa_interes") or solicitud.get("tasa_interes_anual")
    if tasa_raw is None:
        tasa_raw = prestamo.tasa_interes or getattr(prestamo.tipo, "tasa_interes_anual", Decimal("0"))
    try:
        tasa_decimal = Decimal(str(tasa_raw))
    except Exception:
        tasa_decimal = Decimal("0")

    plan = calcular_tabla_amortizacion(prestamo.monto, tasa_decimal, plazo_meses)
    cuota_decimal = Decimal(plan["cuota_mensual"])
    return {
        "plazo_meses": plazo_meses,
        "cuota_mensual": plan["cuota_mensual"],
        "cuota_decimal": cuota_decimal,
    }


def _crear_prestamo_desde_solicitud_row(solicitud_row: dict):
    """
    Genera un registro en prestamo usando los datos de la solicitud aprobada.
    Devuelve (prestamo, error:str|None). Usa el id de la solicitud para mantener trazabilidad/idempotencia.
    """
    solicitud_id = solicitud_row.get("id")
    socio_id = solicitud_row.get("socio_id")
    if not socio_id:
        return None, "La solicitud no tiene socio asociado."
    socio = Socio.objects.filter(pk=socio_id).first()
    if not socio:
        return None, "No se encontró el socio asociado a la solicitud."

    # Evita duplicar si ya existe el préstamo con ese id
    if solicitud_id and Prestamo.objects.filter(pk=solicitud_id).exists():
        return Prestamo.objects.get(pk=solicitud_id), None

    tipo_id = solicitud_row.get("tipo_prestamo_id") or solicitud_row.get("producto_id")
    tipo = TipoPrestamo.objects.filter(pk=tipo_id).first() if tipo_id else None

    try:
        monto = Decimal(str(solicitud_row.get("monto") or "0"))
    except Exception:
        return None, "Monto de solicitud inválido."

    try:
        tasa = Decimal(str(solicitud_row.get("tasa_interes") or "0"))
    except Exception:
        tasa = Decimal("0")

    plazo_raw = solicitud_row.get("plazo_meses")
    try:
        plazo_meses = int(plazo_raw) if plazo_raw is not None else None
    except (TypeError, ValueError):
        plazo_meses = None

    fecha_desembolso = timezone.now().date()
    fecha_vencimiento = add_months(fecha_desembolso, plazo_meses) if plazo_meses else None

    prestamo = Prestamo.objects.create(
        id=solicitud_id or uuid.uuid4(),
        socio=socio,
        tipo=tipo,
        monto=monto,
        tasa_interes=tasa,
        estado="aprobado",
        fecha_desembolso=fecha_desembolso,
        fecha_vencimiento=fecha_vencimiento,
        descripcion=solicitud_row.get("descripcion") or "",
    )
    return prestamo, None


def _desembolso_columnas() -> dict:
    cols = get_table_columns("desembolso")
    return {
        "cols": cols,
        "metodo": "metodo_pago" if "metodo_pago" in cols else ("metodo" if "metodo" in cols else None),
        "fecha": "created_at" if "created_at" in cols else ("fecha" if "fecha" in cols else None),
        "comentarios": "comentarios" if "comentarios" in cols else None,
        "socio": "socio_id" if "socio_id" in cols else None,
        "tesorero": "tesorero_id" if "tesorero_id" in cols else None,
        "updated": "updated_at" if "updated_at" in cols else None,
        "referencia": "referencia" if "referencia" in cols else None,
    }


def _desembolso_prefetch() -> tuple[Prefetch | None, dict]:
    meta = _desembolso_columnas()
    if not meta["cols"]:
        return None, meta
    qs = Desembolso.objects.all()
    if not meta["comentarios"]:
        qs = qs.defer("comentarios")
    return Prefetch("desembolsos", queryset=qs), meta


def _fmt_uuid(val):
    if isinstance(val, uuid.UUID):
        return str(val)
    if isinstance(val, str) and len(val) == 32 and val.count("-") == 0:
        return f"{val[0:8]}-{val[8:12]}-{val[12:16]}-{val[16:20]}-{val[20:]}"
    return val


def _recomendacion_basica(socio: Socio | None, solicitud: dict) -> str:
    if not socio or socio.estado != Socio.ESTADO_ACTIVO:
        return "rechazar"
    try:
        monto = Decimal(str(solicitud.get("monto", "0")))
    except Exception:
        monto = Decimal("0")
    if monto <= Decimal("10000000"):
        return "aprobar"
    if monto <= Decimal("20000000"):
        return "revisar"
    return "rechazar"


class SolicitudEvaluarView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_analista(self, request):
        if not _is_analista(request.user):
            return Response({"detail": "Solo analistas o administradores pueden evaluar solicitudes."}, status=status.HTTP_403_FORBIDDEN)
        return None

    @extend_schema(
        tags=["Prestamos"],
        summary="Evaluar solicitud",
        description="Retorna detalles de la solicitud y socio para ser evaluada por un analista.",
        responses={200: OpenApiResponse(description="Detalle de solicitud para evaluar")},
    )
    def get(self, request, solicitud_id: uuid.UUID):
        forbidden = self._ensure_analista(request)
        if forbidden:
            return forbidden
        solicitud, columnas = _fetch_solicitud_row(solicitud_id)
        socio = Socio.objects.filter(pk=solicitud.get("socio_id")).select_related("usuario").first()
        recomendacion = _recomendacion_basica(socio, solicitud)
        observaciones = _extraer_observaciones(solicitud)

        data = {
            "solicitud": {
                "id": str(solicitud_id),
                "estado": solicitud.get("estado"),
                "monto": str(solicitud.get("monto")),
                "plazo_meses": solicitud.get("plazo_meses"),
                "descripcion": solicitud.get("descripcion"),
                "observaciones": observaciones,
            },
            "socio": None,
            "analisis": {
                "recomendacion": recomendacion,
                "puede_aprobar": recomendacion == "aprobar",
                "columns": list(columnas),
            },
        }
        if socio:
            data["socio"] = {
                "id": str(socio.id),
                "nombre_completo": socio.nombre_completo,
                "documento": socio.documento,
                "estado": socio.estado,
                "email": socio.usuario.email if socio.usuario else None,
                "fecha_alta": socio.fecha_alta,
            }
        return Response(data, status=status.HTTP_200_OK)

    @extend_schema(
        tags=["Prestamos"],
        summary="Guardar observaciones de analista",
        description="Permite registrar observaciones y recomendacion sobre una solicitud.",
        request=None,
        responses={200: OpenApiResponse(description="Observacion registrada")},
    )
    def put(self, request, solicitud_id: uuid.UUID):
        forbidden = self._ensure_analista(request)
        if forbidden:
            return forbidden
        observaciones = (request.data or {}).get("observaciones")
        recomendacion_req = (request.data or {}).get("recomendacion")
        if observaciones is None and recomendacion_req is None:
            return Response({"detail": "Debe enviar observaciones o recomendacion."}, status=status.HTTP_400_BAD_REQUEST)

        solicitud, columnas = _fetch_solicitud_row(solicitud_id)
        obs_col = _obs_column(columnas)
        updates = {}
        descripcion_actual = solicitud.get("descripcion") or ""
        now = timezone.now()
        if observaciones:
            if obs_col:
                updates[obs_col] = observaciones
            else:
                updates["descripcion"] = _adjuntar_observacion_en_descripcion(descripcion_actual, observaciones)
        if recomendacion_req:
            updates["recomendacion"] = recomendacion_req  # solo se guarda si la columna existe
        has_updated = "updated_at" in columnas
        if has_updated:
            updates["updated_at"] = now

        set_parts = []
        values = []
        for col, val in updates.items():
            if col not in columnas and not (col == "updated_at" and has_updated):
                continue
            set_parts.append(f"{col} = %s")
            values.append(val)

        if has_updated and all(not part.startswith("updated_at") for part in set_parts):
            set_parts.append("updated_at = %s")
            values.append(now)

        if not set_parts:
            return Response({"detail": "No hay columnas para actualizar observaciones."}, status=status.HTTP_400_BAD_REQUEST)

        values.append(str(solicitud_id))
        table_name = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"
        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE {table_name} SET {', '.join(set_parts)} WHERE id = %s",
                values,
            )

        solicitud_actualizada, _ = _fetch_solicitud_row(solicitud_id)
        return Response(
            {
                "solicitud_id": str(solicitud_id),
                "estado": solicitud_actualizada.get("estado"),
                "observaciones": observaciones or _extraer_observaciones(solicitud_actualizada),
            },
            status=status.HTTP_200_OK,
        )


class SolicitudDecisionBaseView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    nuevo_estado: str = ""

    def _ensure_analista(self, request):
        if not _is_analista(request.user):
            return Response({"detail": "Solo analistas o administradores pueden decidir solicitudes."}, status=status.HTTP_403_FORBIDDEN)
        return None

    def patch(self, request, solicitud_id: uuid.UUID):
        forbidden = self._ensure_analista(request)
        if forbidden:
            return forbidden
        comentario = (request.data or {}).get("comentario") or ""

        solicitud, columnas = _fetch_solicitud_row(solicitud_id)
        if "estado" not in columnas:
            return Response({"detail": "La tabla de solicitudes no tiene columna 'estado'."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        descripcion_actual = solicitud.get("descripcion") or ""
        now = timezone.now()

        has_updated = "updated_at" in columnas
        updates = {"estado": self.nuevo_estado}
        if has_updated:
            updates["updated_at"] = now
        obs_col = _obs_column(columnas)
        if obs_col and comentario:
            updates[obs_col] = comentario
        elif comentario:
            updates["descripcion"] = _adjuntar_observacion_en_descripcion(descripcion_actual, comentario, marker="[DECISION]")

        set_parts = []
        values = []
        for col, val in updates.items():
            if col not in columnas and not (col == "updated_at" and has_updated):
                continue
            set_parts.append(f"{col} = %s")
            values.append(val)

        values.append(str(solicitud_id))
        table_name = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"
        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(
                    f"UPDATE {table_name} SET {', '.join(set_parts)} WHERE id = %s",
                    values,
                )

            solicitud_actualizada, _ = _fetch_solicitud_row(solicitud_id)
            prestamo_id = None
            if self.nuevo_estado == "aprobado":
                prestamo, error = _crear_prestamo_desde_solicitud_row(solicitud_actualizada)
                if error:
                    raise ValidationError({"detail": error})
                if prestamo:
                    prestamo_id = str(prestamo.id)

        socio = Socio.objects.filter(pk=solicitud_actualizada.get("socio_id")).first()
        notificacion = None
        if socio and getattr(socio, "usuario", None) and socio.usuario.email:
            notificacion = f"Notificar a {socio.usuario.email} sobre estado {self.nuevo_estado}"

        return Response(
            {
                "solicitud_id": str(solicitud_id),
                "estado": solicitud_actualizada.get("estado"),
                "comentario": comentario,
                "notificacion": notificacion,
                "prestamo_id": prestamo_id,
            },
            status=status.HTTP_200_OK,
        )


class SolicitudAprobarView(SolicitudDecisionBaseView):
    nuevo_estado = "aprobado"


class SolicitudRechazarView(SolicitudDecisionBaseView):
    nuevo_estado = "rechazado"


class SolicitudListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not _is_analista(request.user):
            return Response({"detail": "Solo analistas o administradores pueden listar solicitudes."}, status=status.HTTP_403_FORBIDDEN)

        q = (request.query_params.get("q") or "").strip().lower()
        estado = (request.query_params.get("estado") or "").strip().lower()
        limit = min(max(int(request.query_params.get("limit", 20)), 1), 100)

        columnas = get_table_columns("solicitud")
        if not columnas:
            return Response({"detail": "Tabla solicitud no disponible."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        base_cols = ["id", "socio_id", "monto", "plazo_meses", "estado", "created_at", "updated_at", "descripcion"]
        select_cols = [c for c in base_cols if c in columnas]
        if "observaciones" in columnas:
            select_cols.append("observaciones")
        table_name = "solicitud" if connection.vendor != "postgresql" else "public.solicitud"

        where_parts = []
        params: list = []
        if estado:
            where_parts.append("LOWER(estado) = %s")
            params.append(estado)
        if q:
            where_parts.append("(LOWER(id::text) LIKE %s OR LOWER(descripcion) LIKE %s)")
            like = f"%{q}%"
            params.extend([like, like])

        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
        sql = f"""
            SELECT {', '.join(select_cols)}
            FROM {table_name}
            {where_sql}
            ORDER BY created_at DESC
            LIMIT %s
        """
        params.append(limit)

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        results = []
        for row in rows:
            data = dict(zip(select_cols, row))
            socio = Socio.objects.filter(pk=data.get("socio_id")).select_related("usuario").first()
            obs_col = _obs_column(columnas)
            results.append(
                {
                    "id": str(data.get("id")),
                    "estado": data.get("estado"),
                    "monto": str(data.get("monto")),
                    "plazo_meses": data.get("plazo_meses"),
                    "descripcion": data.get("descripcion"),
                    "observaciones": data.get(obs_col) if obs_col and obs_col in data else _extraer_observaciones(data),
                    "created_at": data.get("created_at"),
                    "socio": {
                        "id": str(socio.id),
                        "nombre_completo": socio.nombre_completo,
                        "documento": socio.documento,
                        "email": socio.usuario.email if socio and socio.usuario else None,
                        "estado": socio.estado if socio else None,
                        "fecha_alta": socio.fecha_alta if socio else None,
                    } if socio else None,
                }
            )
        return Response({"results": results, "count": len(results)}, status=status.HTTP_200_OK)


class PrestamosAprobadosListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_tesorero(self, request):
        if not _is_tesorero(request.user):
            return Response({"detail": "Solo tesorero o admin puede ver préstamos aprobados."}, status=status.HTTP_403_FORBIDDEN)
        return None

    def get(self, request):
        forbidden = self._ensure_tesorero(request)
        if forbidden:
            return forbidden

        q = (request.query_params.get("q") or "").strip()
        limit = min(max(int(request.query_params.get("limit", 30)), 1), 100)
        estados_permitidos = {"aprobado", Prestamo.Estados.ACTIVO, "activo"}
        qs = (
            Prestamo.objects.select_related("socio", "socio__usuario")
            .annotate(
                desembolsos_count=Count("desembolsos"),
                pagos_count=Count("pagos"),
                estado_norm=Lower(Trim("estado")),
            )
            .filter(estado_norm__in=["aprobado", "activo"])
            .filter(desembolsos_count=0, pagos_count=0)
            .order_by("-created_at")
        )
        if q:
            qs = qs.filter(
                Q(id__icontains=q)
                | Q(descripcion__icontains=q)
                | Q(socio__nombre_completo__icontains=q)
                | Q(socio__documento__icontains=q)
            )
        qs = qs[:limit]

        results = []
        for prestamo in qs:
            socio = prestamo.socio
            results.append(
                {
                    "id": str(prestamo.id),
                    "monto": fmt_decimal(prestamo.monto) if hasattr(prestamo.monto, "quantize") else str(prestamo.monto),
                    "estado": prestamo.estado,
                    "descripcion": prestamo.descripcion,
                    "socio": {
                        "id": str(socio.id),
                        "nombre_completo": socio.nombre_completo,
                        "documento": socio.documento,
                        "email": socio.usuario.email if socio and socio.usuario else None,
                    }
                    if socio
                    else None,
                }
            )

        return Response({"results": results, "count": len(results)}, status=status.HTTP_200_OK)


class DesembolsoListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_tesorero(self, request):
        if not _is_tesorero(request.user):
            return Response({"detail": "Solo tesorero o admin puede gestionar desembolsos."}, status=status.HTTP_403_FORBIDDEN)
        return None

    def get(self, request):
        forbidden = self._ensure_tesorero(request)
        if forbidden:
            return forbidden
        meta = _desembolso_columnas()
        cols = ["id", "prestamo_id", "monto"]
        if meta["metodo"]:
            cols.append(meta["metodo"])
        if meta["referencia"]:
            cols.append(meta["referencia"])
        if meta["comentarios"]:
            cols.append(meta["comentarios"])
        if meta["fecha"]:
            cols.append(meta["fecha"])
        if meta["updated"]:
            cols.append(meta["updated"])
        socio_join = meta["socio"] is not None
        socio_cols = []
        if socio_join:
            socio_cols = ["nombre_completo", "documento", "email"]
        select_cols = ", ".join([f"d.{c}" for c in cols])
        table_name = "desembolso" if connection.vendor != "postgresql" else "public.desembolso"
        sql = f"SELECT {select_cols}"
        if socio_join:
            sql += ", s.nombre_completo, s.documento, u.email"
        sql += f" FROM {table_name} d"
        if socio_join:
            sql += " LEFT JOIN socio s ON s.id = d.socio_id LEFT JOIN usuario u ON u.id = s.usuario_id"
        sql += " ORDER BY d.id DESC LIMIT 100"

        data = []
        with connection.cursor() as cursor:
            cursor.execute(sql)
            rows = cursor.fetchall()
        for row in rows:
            base = dict(zip([c.split(".")[-1] for c in cols], row[: len(cols)]))
            record = {
                "id": str(_fmt_uuid(base.get("id"))),
                "prestamo_id": str(_fmt_uuid(base.get("prestamo_id"))),
                "monto": str(base.get("monto")),
                "metodo_pago": base.get(meta["metodo"]) if meta["metodo"] else None,
                "referencia": base.get(meta["referencia"]) if meta["referencia"] else None,
                "comentarios": base.get(meta["comentarios"]) if meta["comentarios"] else None,
                "created_at": base.get(meta["fecha"]) if meta["fecha"] else None,
                "updated_at": base.get(meta["updated"]) if meta["updated"] else None,
                "socio": None,
            }
            if socio_join:
                socio_data = row[len(cols) :]
                record["socio"] = {
                    "nombre_completo": socio_data[0],
                    "documento": socio_data[1],
                    "email": socio_data[2],
                }
            data.append(record)
        return Response({"results": data, "count": len(data)}, status=status.HTTP_200_OK)

    def post(self, request):
        forbidden = self._ensure_tesorero(request)
        if forbidden:
            return forbidden
        meta = _desembolso_columnas()
        prestamo_id = (request.data or {}).get("prestamo_id")
        monto = (request.data or {}).get("monto")
        metodo_pago = (request.data or {}).get("metodo_pago") or (request.data or {}).get("metodo")
        referencia = (request.data or {}).get("referencia") or ""
        comentarios = (request.data or {}).get("comentarios") or ""

        if not prestamo_id or not monto or not metodo_pago:
            return Response({"detail": "prestamo_id, monto y metodo_pago son obligatorios"}, status=status.HTTP_400_BAD_REQUEST)

        prestamo = get_object_or_404(Prestamo, pk=prestamo_id)
        estado = (prestamo.estado or "").lower()
        if estado not in {"aprobado", Prestamo.Estados.ACTIVO, "activo"}:
            return Response({"prestamo_id": "El préstamo no está aprobado/activo para desembolso."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            monto_decimal = Decimal(str(monto))
        except Exception:
            return Response({"monto": "Monto inválido."}, status=status.HTTP_400_BAD_REQUEST)
        if prestamo.monto and monto_decimal > prestamo.monto:
            return Response({"monto": "El monto excede el valor del préstamo."}, status=status.HTTP_400_BAD_REQUEST)

        # Si la tabla coincide con el esquema del modelo (metodo_pago/created_at y columnas clave), usamos ORM
        can_use_orm = (
            meta["metodo"] == "metodo_pago"
            and meta["fecha"] == "created_at"
            and meta["comentarios"] == "comentarios"
            and meta["socio"] == "socio_id"
        )
        if can_use_orm:
            serializer = DesembolsoSerializer(data={
                "prestamo_id": prestamo.id,
                "monto": monto_decimal,
                "metodo_pago": metodo_pago,
                "referencia": referencia,
                "comentarios": comentarios,
            })
            serializer.is_valid(raise_exception=True)
            desembolso = serializer.save()

            if prestamo.estado not in {Prestamo.Estados.PAGADO, Prestamo.Estados.CANCELADO}:
                prestamo.estado = "desembolsado"
                prestamo.save(update_fields=["estado", "updated_at"])

            return Response(DesembolsoSerializer(desembolso).data, status=status.HTTP_201_CREATED)

        payload = {
            "id": uuid.uuid4(),
            "prestamo_id": prestamo.id,
            "monto": monto_decimal,
        }
        if meta["metodo"]:
            payload[meta["metodo"]] = metodo_pago
        if meta["referencia"]:
            payload[meta["referencia"]] = referencia
        if meta["comentarios"]:
            payload[meta["comentarios"]] = comentarios
        if meta["fecha"]:
            payload[meta["fecha"]] = timezone.now()
        if meta["updated"]:
            payload[meta["updated"]] = timezone.now()
        if meta["socio"]:
            payload[meta["socio"]] = prestamo.socio_id
        if meta.get("tesorero"):
            payload[meta["tesorero"]] = getattr(request.user, "id", None)

        cols_presentes = [c for c in payload.keys() if c in meta["cols"] or c in {meta.get("fecha"), meta.get("metodo"), meta.get("referencia"), meta.get("comentarios"), meta.get("updated"), meta.get("socio")}]
        placeholders = ", ".join(["%s"] * len(cols_presentes))
        columnas_sql = ", ".join(cols_presentes)
        valores = [payload[col] for col in cols_presentes]
        if connection.vendor == "sqlite":
            conv = []
            for v in valores:
                if isinstance(v, uuid.UUID):
                    conv.append(str(v))
                elif isinstance(v, Decimal):
                    conv.append(float(v))
                else:
                    conv.append(v)
            valores = conv
        table_name = "desembolso" if connection.vendor != "postgresql" else "public.desembolso"
        with connection.cursor() as cursor:
            cursor.execute(
                f"INSERT INTO {table_name} ({columnas_sql}) VALUES ({placeholders})",
                valores,
            )

        if prestamo.estado not in {Prestamo.Estados.PAGADO, Prestamo.Estados.CANCELADO}:
            prestamo.estado = "desembolsado"
            prestamo.save(update_fields=["estado", "updated_at"])

        return Response(
            {
                "id": str(payload["id"]),
                "prestamo_id": str(prestamo.id),
                "monto": str(monto_decimal),
                "metodo_pago": metodo_pago,
                "referencia": referencia,
                "comentarios": comentarios if meta["comentarios"] else None,
                "socio": {
                    "id": str(prestamo.socio_id),
                    "nombre_completo": prestamo.socio.nombre_completo if prestamo.socio else None,
                    "documento": prestamo.socio.documento if prestamo.socio else None,
                    "email": prestamo.socio.usuario.email if prestamo.socio and prestamo.socio.usuario else None,
                } if meta["socio"] else None,
            },
            status=status.HTTP_201_CREATED,
        )


class PoliticaAprobacionListCreateView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Configuracion'],
        responses=PoliticaAprobacionSerializer(many=True),
        summary='Listado de políticas de aprobación',
        description='Devuelve las políticas de aprobación automática configuradas. Solo administradores.',
    )
    def get(self, request):
        qs = PoliticaAprobacion.objects.all().order_by('nombre')
        q = (request.query_params.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))

        solo_activas = request.query_params.get('solo_activos') or request.query_params.get('soloActivos')
        if solo_activas and str(solo_activas).lower() in {'1', 'true', 't', 'yes', 'on'}:
            qs = qs.filter(activo=True)

        return Response(PoliticaAprobacionSerializer(qs, many=True).data)

    @extend_schema(
        tags=['Configuracion'],
        request=PoliticaAprobacionUpsertSerializer,
        responses={201: PoliticaAprobacionSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Crear política de aprobación',
        description='Registra una política basada en score, antigüedad y capacidad de pago.',
    )
    def post(self, request):
        serializer = PoliticaAprobacionUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        politica = serializer.save()
        return Response(PoliticaAprobacionSerializer(politica).data, status=status.HTTP_201_CREATED)


class PoliticaAprobacionDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get_object(self, politica_id):
        return get_object_or_404(PoliticaAprobacion, pk=politica_id)

    @extend_schema(
        tags=['Configuracion'],
        responses={200: PoliticaAprobacionSerializer, 404: OpenApiResponse(description='Política no encontrada')},
        summary='Detalle de política de aprobación',
        description='Obtiene la definición de una política de aprobación automática.',
    )
    def get(self, _request, politica_id):
        politica = self.get_object(politica_id)
        return Response(PoliticaAprobacionSerializer(politica).data)

    @extend_schema(
        tags=['Configuracion'],
        request=PoliticaAprobacionUpsertSerializer,
        responses={200: PoliticaAprobacionSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Actualizar política de aprobación',
        description='Actualiza los parámetros de score, antigüedad y capacidad de pago.',
    )
    def put(self, request, politica_id):
        politica = self.get_object(politica_id)
        serializer = PoliticaAprobacionUpsertSerializer(instance=politica, data=request.data)
        serializer.is_valid(raise_exception=True)
        politica = serializer.save()
        return Response(PoliticaAprobacionSerializer(politica).data)

    @extend_schema(
        tags=['Configuracion'],
        request=PoliticaAprobacionUpsertSerializer,
        responses={200: PoliticaAprobacionSerializer, 400: OpenApiResponse(description='Datos inválidos')},
        summary='Actualización parcial de política',
        description='Permite editar parcialmente una política de aprobación.',
    )
    def patch(self, request, politica_id):
        politica = self.get_object(politica_id)
        serializer = PoliticaAprobacionUpsertSerializer(instance=politica, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        politica = serializer.save()
        return Response(PoliticaAprobacionSerializer(politica).data)

    @extend_schema(
        tags=['Configuracion'],
        responses={200: PoliticaAprobacionSerializer, 404: OpenApiResponse(description='Política no encontrada')},
        summary='Desactivar política de aprobación',
        description='Desactiva (soft delete) una política para que no aplique a nuevas solicitudes.',
    )
    def delete(self, _request, politica_id):
        politica = self.get_object(politica_id)
        if politica.activo:
            politica.activo = False
            politica.save(update_fields=['activo', 'updated_at'])
        return Response(PoliticaAprobacionSerializer(politica).data, status=status.HTTP_200_OK)


class PoliticaAprobacionPublicListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not (_is_analista(request.user) or getattr(request.user, "is_staff", False)):
            return Response({"detail": "No autorizado"}, status=status.HTTP_403_FORBIDDEN)
        qs = PoliticaAprobacion.objects.filter(activo=True).order_by("nombre")
        return Response(PoliticaAprobacionSerializer(qs, many=True).data)


class SocioAdminDetailView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get_object(self, socio_id):
        return get_object_or_404(Socio.objects.select_related('usuario'), pk=socio_id)

    @extend_schema(
        tags=['Socios'],
        responses=SocioSerializer,
        summary='Detalle socio',
        description='Obtiene la información completa de un socio específico',
    )
    def get(self, _request, socio_id):
        socio = self.get_object(socio_id)
        return Response(SocioSerializer(socio).data)

    @extend_schema(
        tags=['Socios'],
        request=SocioAdminUpdateSerializer,
        responses={
            200: SocioSerializer,
            400: OpenApiResponse(description='Errores de validación'),
            404: OpenApiResponse(description='Socio no encontrado'),
        },
        summary='Actualización completa',
        description='Actualiza campos permitidos del socio. No permite modificar estado o campos críticos.',
    )
    def put(self, request, socio_id):
        socio = self.get_object(socio_id)
        serializer = SocioAdminUpdateSerializer(instance=socio, data=request.data)
        serializer.is_valid(raise_exception=True)
        before = snapshot_socio(socio)
        socio = serializer.save()
        after = snapshot_socio(socio)
        register_audit_entry(
            socio=socio,
            user=request.user,
            action=SocioAuditLog.Actions.UPDATE,
            before=before,
            after=after,
        )
        return Response(SocioSerializer(socio).data)

    @extend_schema(
        tags=['Socios'],
        responses={
            200: SocioSerializer,
            404: OpenApiResponse(description='Socio no encontrado'),
        },
        summary='Baja lógica de socio',
        description='Marca al socio como inactivo (baja lógica). Solo administradores.',
    )
    def delete(self, _request, socio_id):
        socio = self.get_object(socio_id)
        before = snapshot_socio(socio)
        if socio.estado != Socio.ESTADO_INACTIVO:
            socio.estado = Socio.ESTADO_INACTIVO
            socio.save(update_fields=['estado', 'updated_at'])
            after = snapshot_socio(socio)
            register_audit_entry(
                socio=socio,
                user=getattr(_request, 'user', None),
                action=SocioAuditLog.Actions.STATE_CHANGE,
                before=before,
                after=after,
                metadata={'motivo': 'baja_logica'},
            )
        return Response(SocioSerializer(socio).data, status=status.HTTP_200_OK)


class SocioEstadoUpdateView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get_object(self, socio_id):
        return get_object_or_404(Socio, pk=socio_id)

    @extend_schema(
        tags=['Socios'],
        request=SocioEstadoSerializer,
        responses={
            200: SocioSerializer,
            400: OpenApiResponse(description='Transición no permitida o datos inválidos'),
            404: OpenApiResponse(description='Socio no encontrado'),
        },
        summary='Cambio de estado',
        description='Cambia el estado del socio controlando transiciones válidas.',
    )
    def patch(self, request, socio_id):
        socio = self.get_object(socio_id)
        serializer = SocioEstadoSerializer(data=request.data, context={'socio': socio})
        serializer.is_valid(raise_exception=True)
        before = snapshot_socio(socio)
        socio.estado = serializer.validated_data['estado']
        socio.save(update_fields=['estado', 'updated_at'])
        after = snapshot_socio(socio)
        register_audit_entry(
            socio=socio,
            user=request.user,
            action=SocioAuditLog.Actions.STATE_CHANGE,
            before=before,
            after=after,
            metadata={'motivo': serializer.validated_data.get('motivo') or ''},
        )
        return Response(SocioSerializer(socio).data)


class SocioHistorialView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Socios'],
        responses={200: HistorialCrediticioSerializer},
        summary='Historial crediticio de un socio',
        description='Devuelve los préstamos previos del socio y sus pagos, filtrables por estado y rango de fechas.',
    )
    def get(self, request, socio_id=None):
        socio = get_object_or_404(Socio, pk=socio_id) if socio_id else None

        estados_param = request.query_params.get('estado') or ''
        estados = {e.strip() for e in estados_param.split(',') if e.strip()}
        estados_validos = set(Prestamo.Estados.values)
        if estados and not estados.issubset(estados_validos):
            desconocidos = estados - estados_validos
            raise ValidationError({'estado': [f"Estado(s) desconocido(s): {', '.join(desconocidos)}"]})

        def parse_date(param_name: str):
            valor = request.query_params.get(param_name)
            if not valor:
                return None
            try:
                return datetime.fromisoformat(valor).date()
            except Exception:
                raise ValidationError({param_name: 'Usa formato ISO AAAA-MM-DD.'})

        desde = parse_date('desde')
        hasta = parse_date('hasta')

        prestamos_qs = Prestamo.objects.filter(socio=socio) if socio else Prestamo.objects.all()
        prestamos_qs = prestamos_qs.select_related('socio', 'tipo').prefetch_related('pagos').order_by('-fecha_desembolso')
        if estados:
            prestamos_qs = prestamos_qs.filter(estado__in=estados)
        if desde:
            prestamos_qs = prestamos_qs.filter(fecha_desembolso__gte=desde)
        if hasta:
            prestamos_qs = prestamos_qs.filter(fecha_desembolso__lte=hasta)

        prestamos = []
        for prestamo in prestamos_qs:
            pagos_qs = prestamo.pagos.all()
            if desde:
                pagos_qs = pagos_qs.filter(fecha_pago__gte=desde)
            if hasta:
                pagos_qs = pagos_qs.filter(fecha_pago__lte=hasta)
            prestamo._pagos_filtrados = list(pagos_qs)
            prestamos.append(prestamo)

        serializer = HistorialCrediticioSerializer({
            'socio': socio,
            'prestamos': prestamos,
        })
        return Response(serializer.data)


class SocioExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Socios'],
        summary='Exportar socios y auditoria a Excel',
        description=(
            'Genera un archivo Excel con los socios (filtrables por estado) '
            'y el historial de auditoria de cambios, filtrable por rango de fechas y accion.'
        ),
        responses={200: OpenApiResponse(description='Excel exportado')},
    )
    def get(self, request):
        estados_param = request.query_params.get('estado')
        estados = {e.strip() for e in estados_param.split(',')} if estados_param else set()
        accion = request.query_params.get('accion') or None
        desde_param = request.query_params.get('desde')
        hasta_param = request.query_params.get('hasta')

        def parse_dt(value: str):
            try:
                parsed = datetime.fromisoformat(value)
                return parsed if parsed.tzinfo else timezone.make_aware(parsed)
            except Exception:
                return None

        desde = parse_dt(desde_param) if desde_param else None
        hasta = parse_dt(hasta_param) if hasta_param else None

        socios_qs = Socio.objects.select_related('usuario').order_by('nombre_completo')
        if estados:
            socios_qs = socios_qs.filter(estado__in=estados)

        audit_qs = SocioAuditLog.objects.select_related('socio', 'performed_by').order_by('-created_at')
        if accion:
            audit_qs = audit_qs.filter(action=accion)
        if desde:
            audit_qs = audit_qs.filter(created_at__gte=desde)
        if hasta:
            audit_qs = audit_qs.filter(created_at__lte=hasta)

        wb = Workbook()

        # Resumen / filtros aplicados
        ws_meta = wb.active
        ws_meta.title = "Resumen"
        now = timezone.localtime()
        ws_meta.append(["Reporte generado"])
        ws_meta.append(["Generado por", getattr(request.user, 'email', '')])
        ws_meta.append(["Fecha/Hora", now.strftime("%Y-%m-%d %H:%M:%S %Z")])
        ws_meta.append(["Filtros", ""])
        ws_meta.append(["  Estado", ", ".join(sorted(estados)) if estados else "Todos"])
        ws_meta.append(["  Accion (auditoria)", accion or "Todas"])
        ws_meta.append(["  Desde", desde.strftime("%Y-%m-%d %H:%M:%S") if desde else ""])
        ws_meta.append(["  Hasta", hasta.strftime("%Y-%m-%d %H:%M:%S") if hasta else ""])
        for cell in ws_meta["A"]:
            cell.font = Font(bold=True)

        # Hoja de socios
        ws_socios = wb.create_sheet("Socios")
        socios_headers = [
            "ID", "Nombre", "Documento", "Estado", "Email usuario",
            "Telefono", "Direccion", "Fecha alta", "Creado", "Actualizado",
            "Usuario ID",
        ]
        ws_socios.append(socios_headers)
        for socio in socios_qs:
            ws_socios.append([
                str(socio.id),
                socio.nombre_completo,
                socio.documento or "",
                socio.estado,
                socio.usuario.email if socio.usuario else "",
                socio.telefono or "",
                socio.direccion or "",
                socio.fecha_alta.isoformat() if socio.fecha_alta else "",
                socio.created_at.isoformat(),
                socio.updated_at.isoformat(),
                str(socio.usuario.id) if socio.usuario else "",
            ])
        style_header_row(ws_socios)
        ws_socios.freeze_panes = "A2"
        ws_socios.auto_filter.ref = f"A1:{get_column_letter(len(socios_headers))}{ws_socios.max_row}"
        for idx in range(1, len(socios_headers) + 1):
            ws_socios.column_dimensions[get_column_letter(idx)].width = 18

        # Hoja de auditoria
        ws_audit = wb.create_sheet("Auditoria")
        audit_headers = [
            "ID", "Socio ID", "Socio", "Email", "Accion",
            "Estado anterior", "Estado nuevo", "Campos modificados",
            "Datos previos", "Datos nuevos", "Metadata",
            "Ejecutado por", "Fecha",
        ]
        ws_audit.append(audit_headers)
        for entry in audit_qs:
            socio = entry.socio
            ws_audit.append([
                entry.id,
                str(socio.id) if socio else "",
                socio.nombre_completo if socio else "",
                socio.usuario.email if socio and socio.usuario else "",
                entry.action,
                entry.estado_anterior,
                entry.estado_nuevo,
                ", ".join(entry.campos_modificados or []),
                json.dumps(entry.datos_previos or {}, ensure_ascii=False),
                json.dumps(entry.datos_nuevos or {}, ensure_ascii=False),
                json.dumps(entry.metadata or {}, ensure_ascii=False),
                entry.performed_by.email if entry.performed_by else "",
                entry.created_at.isoformat(),
            ])
        style_header_row(ws_audit)
        ws_audit.freeze_panes = "A2"
        ws_audit.auto_filter.ref = f"A1:{get_column_letter(len(audit_headers))}{ws_audit.max_row}"
        wrap_columns(ws_audit, ["H", "I", "J", "K"])
        for idx in range(1, len(audit_headers) + 1):
            ws_audit.column_dimensions[get_column_letter(idx)].width = 20

        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"socios_auditoria_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class SocioHistorialExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Socios'],
        summary='Exportar historial crediticio',
        description='Genera un XLSX con los préstamos y pagos del socio (o todos) aplicando los filtros.',
        responses={200: OpenApiResponse(description='Excel exportado')},
    )
    def get(self, request, socio_id=None):
        estados_param = request.query_params.get('estado') or ''
        estados = {e.strip() for e in estados_param.split(',') if e.strip()}
        estados_validos = set(Prestamo.Estados.values)
        if estados and not estados.issubset(estados_validos):
            desconocidos = estados - estados_validos
            raise ValidationError({'estado': [f"Estado(s) desconocido(s): {', '.join(desconocidos)}"]})

        def parse_date(param_name: str):
            valor = request.query_params.get(param_name)
            if not valor:
                return None
            try:
                return datetime.fromisoformat(valor).date()
            except Exception:
                raise ValidationError({param_name: 'Usa formato ISO AAAA-MM-DD.'})

        desde = parse_date('desde')
        hasta = parse_date('hasta')

        socio = None
        if socio_id:
            socio = get_object_or_404(Socio, pk=socio_id)

        prestamos_qs = Prestamo.objects.select_related('socio', 'socio__usuario', 'tipo').prefetch_related('pagos')
        if socio:
            prestamos_qs = prestamos_qs.filter(socio=socio)
        if estados:
            prestamos_qs = prestamos_qs.filter(estado__in=estados)
        if desde:
            prestamos_qs = prestamos_qs.filter(fecha_desembolso__gte=desde)
        if hasta:
            prestamos_qs = prestamos_qs.filter(fecha_desembolso__lte=hasta)

        wb = Workbook()

        ws_meta = wb.active
        ws_meta.title = "Resumen"
        now = timezone.localtime()
        ws_meta.append(["Reporte generado"])
        ws_meta.append(["Generado por", getattr(request.user, 'email', '')])
        ws_meta.append(["Fecha/Hora", now.strftime("%Y-%m-%d %H:%M:%S %Z")])
        ws_meta.append(["Socio", socio.nombre_completo if socio else "Todos"])
        ws_meta.append(["Filtros", ""])
        ws_meta.append(["  Estado", ", ".join(sorted(estados)) if estados else "Todos"])
        ws_meta.append(["  Desde", desde.strftime("%Y-%m-%d") if desde else ""])
        ws_meta.append(["  Hasta", hasta.strftime("%Y-%m-%d") if hasta else ""])
        for cell in ws_meta["A"]:
            cell.font = Font(bold=True)

        # Hoja de prestamos
        ws_prestamos = wb.create_sheet("Prestamos")
        headers_prestamos = [
            "ID", "Tipo", "Tasa anual", "Plazo (meses)", "Socio", "Documento", "Estado",
            "Monto", "Pagado", "Saldo", "Monto en mora",
            "Días mora", "Cuotas vencidas",
            "Desembolso", "Vencimiento", "Descripción",
        ]
        ws_prestamos.append(headers_prestamos)
        for p in prestamos_qs:
            total_pagado = sum((pg.monto for pg in p.pagos.all()), 0)
            saldo = p.monto - total_pagado
            saldo = saldo if saldo > 0 else 0
            dias_mora = 0
            if p.fecha_vencimiento:
                hoy = date.today()
                if p.fecha_vencimiento < hoy and saldo > 0:
                    dias_mora = (hoy - p.fecha_vencimiento).days
            cuotas_vencidas = max(1, (dias_mora + 29) // 30) if dias_mora > 0 else 0
            monto_mora = saldo if dias_mora > 0 else 0
            socio_name = p.socio.nombre_completo if p.socio else ""
            socio_doc = p.socio.documento if p.socio else ""
            tipo_nombre = p.tipo.nombre if p.tipo else ""
            tipo_tasa = float(p.tipo.tasa_interes_anual) if p.tipo else None
            tipo_plazo = p.tipo.plazo_meses if p.tipo else None

            ws_prestamos.append([
                str(p.id),
                tipo_nombre,
                tipo_tasa,
                tipo_plazo,
                socio_name,
                socio_doc,
                p.estado,
                float(p.monto),
                float(total_pagado),
                float(saldo),
                float(monto_mora),
                dias_mora,
                cuotas_vencidas,
                p.fecha_desembolso.isoformat(),
                p.fecha_vencimiento.isoformat() if p.fecha_vencimiento else "",
                p.descripcion,
            ])
        style_header_row(ws_prestamos)
        ws_prestamos.freeze_panes = "A2"

        # Hoja de pagos
        ws_pagos = wb.create_sheet("Pagos")
        headers_pagos = ["ID", "Préstamo ID", "Socio", "Monto", "Método", "Fecha", "Referencia"]
        ws_pagos.append(headers_pagos)
        for p in prestamos_qs:
            for pago in p.pagos.all():
                ws_pagos.append([
                  pago.id,
                  str(p.id),
                  p.socio.nombre_completo if p.socio else "",
                  float(pago.monto),
                  pago.metodo,
                  pago.fecha_pago.isoformat(),
                  pago.referencia,
                ])
        style_header_row(ws_pagos)
        ws_pagos.freeze_panes = "A2"

        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename_base = "historial_crediticio"
        if socio:
            filename_base = f"historial_{socio.id}"
        filename = f"{filename_base}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class AdminActivityView(APIView):
    permission_classes = [permissions.IsAdminUser]

    @extend_schema(
        tags=['Socios'],
        summary='Actividad reciente del admin',
        description='Últimas acciones del administrador autenticado sobre socios.',
        responses={200: OpenApiResponse(description='Listado de actividad')},
    )
    def get(self, request):
        usuario = request.user
        qs = (
            SocioAuditLog.objects.select_related('socio')
            .filter(performed_by=usuario)
            .order_by('-created_at')[:10]
        )
        data = [
            {
                "socio": log.socio.nombre_completo if log.socio else "",
                "accion": log.get_action_display(),
                "estado_anterior": log.estado_anterior,
                "estado_nuevo": log.estado_nuevo,
                "campos": log.campos_modificados,
                "fecha": log.created_at.isoformat(),
            }
            for log in qs
        ]
        return Response(data)
